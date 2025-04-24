import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# 엑셀 파일 읽기
df = pd.read_excel('app/analysis/output/ssafy_result_0408.xlsx')

# 데이터 확인
print("데이터 샘플:")
print(df.head())

# 요일 매핑 생성
day_mapping = {
    1: '월요일', 2: '월요일',
    3: '화요일', 4: '화요일', 
    5: '수요일', 6: '수요일',
    7: '목요일', 8: '목요일',
    9: '금요일', 10: '금요일'
}

# 각 Time에 대한 요일 추가
df['Day'] = df['Time'].map(day_mapping)

# Line 목록 가져오기
lines = sorted(df['Line'].unique())

# 새 엑셀 파일 생성
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "주간생산계획표"

# 요일 및 Time 헤더 설정
sheet.cell(row=1, column=1).value = "Line"
sheet.cell(row=2, column=1).value = ""

# 요일 헤더 (각 요일은 해당 요일의 Time 수만큼 열을 차지)
days = ['월요일', '화요일', '수요일', '목요일', '금요일']
day_times = {
    '월요일': [1, 2],
    '화요일': [3, 4],
    '수요일': [5, 6],
    '목요일': [7, 8],
    '금요일': [9, 10]
}

current_col = 2
for day in days:
    # 해당 요일의 Time 개수 확인
    times = day_times[day]
    
    # 요일 헤더 병합
    if len(times) > 0:
        sheet.cell(row=1, column=current_col).value = day
        if len(times) > 1:
            sheet.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(times) - 1)
        sheet.cell(row=1, column=current_col).alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=current_col).font = Font(bold=True)
        
        # 각 Time 헤더 추가
        for i, time in enumerate(times):
            sheet.cell(row=2, column=current_col + i).value = f"Time {time}"
            sheet.cell(row=2, column=current_col + i).alignment = Alignment(horizontal='center')
        
        current_col += len(times)

# 각 Line별 데이터 추가
for i, line in enumerate(lines):
    row_num = i + 3  # 헤더가 2행이므로 데이터는 3행부터
    
    # Line 입력
    sheet.cell(row=row_num, column=1).value = line
    
    # 각 요일별 Time별 데이터 입력
    current_col = 2
    for day in days:
        times = day_times[day]
        
        for i, time in enumerate(times):
            # 해당 Line과 Time에 대한 데이터 필터링
            time_data = df[(df['Line'] == line) & (df['Time'] == time)]
            
            if not time_data.empty:
                # 제품 정보와 수량 결합
                items_info = []
                for _, row in time_data.iterrows():
                    item_info = f"{row['Demand']} ({row['Qty']}개)"
                    items_info.append(item_info)
                
                # 셀에 데이터 입력
                sheet.cell(row=row_num, column=current_col + i).value = "\n".join(items_info)
                sheet.cell(row=row_num, column=current_col + i).alignment = Alignment(wrap_text=True)
        
        current_col += len(times)

# 열 너비 조정
sheet.column_dimensions[get_column_letter(1)].width = 10  # Line
for col in range(2, current_col):
    sheet.column_dimensions[get_column_letter(col)].width = 25  # Time 열

# 행 높이 조정
for row in range(1, len(lines) + 3):
    sheet.row_dimensions[row].height = 30

# 테두리 스타일 설정
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# 셀 테두리 적용
for row in range(1, len(lines) + 3):
    for col in range(1, current_col):
        sheet.cell(row=row, column=col).border = thin_border

# 파일 저장
workbook.save('weekly_production_schedule_horizontal.xlsx')
print("\n결과가 'weekly_production_schedule_horizontal.xlsx' 파일로 저장되었습니다.")